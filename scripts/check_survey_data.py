#!/usr/bin/env python3
# pyright: reportMissingImports=false, reportMissingModuleSource=false
"""CLI to check survey data presence/missing counts by module/table/health center."""

from __future__ import annotations

import argparse
import datetime as dt
import difflib
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast

import psycopg
from psycopg import sql
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from module_config import (
    DEFAULT_PROJECT_IDS,
    MODULES,
    PROJECT_DATA_TABLES,
    TABLE_EXPECTED_RECORDS,
    TABLE_NAME_OVERRIDES,
)


@dataclass
class TableResolution:
    module: str
    requested_name: str
    actual_table: str | None
    status: str
    link_column: str | None
    expected_per_survey: int
    notes: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Check survey data presence/missing counts by module, table, and health center.",
    )
    parser.add_argument("--env-file", default=".env", help="Path to env file with DB_* values")
    parser.add_argument("--district", help="District name filter (exact match)")
    parser.add_argument("--district-id", help="District id filter")
    parser.add_argument("--year", help="Survey year filter, e.g. 2024")
    parser.add_argument(
        "--type-ids",
        help="Comma-separated qr_code_types.id values to filter surveys",
    )
    parser.add_argument(
        "--projects",
        help="Comma-separated list of project IDs for project section",
    )
    parser.add_argument(
        "--output-dir",
        default="reports",
        help="Directory where run artifacts are created",
    )
    parser.add_argument(
        "--no-interactive",
        action="store_true",
        help="Disable prompts; use provided flags and defaults",
    )
    return parser.parse_args()


def load_env(path: str) -> dict[str, str]:
    env: dict[str, str] = {}
    env_path = Path(path)
    if not env_path.exists():
        raise FileNotFoundError(f"Env file not found: {path}")

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")

    return env


def db_dsn(env: dict[str, str]) -> str:
    return (
        f"host={env.get('DB_HOST', '')} "
        f"port={env.get('DB_PORT', '5432')} "
        f"dbname={env.get('DB_NAME', '')} "
        f"user={env.get('DB_USER', '')} "
        f"password={env.get('DB_PASSWORD', '')} "
        f"sslmode={env.get('DB_SSLMODE', 'prefer')}"
    )


def normalize_name(name: str) -> str:
    if not name:
        return ""
    if "_" in name:
        return re.sub(r"[^a-z0-9_]", "", name.lower())
    s1 = re.sub("(.)([A-Z][a-z]+)", r"\1_\2", name)
    s2 = re.sub("([a-z0-9])([A-Z])", r"\1_\2", s1)
    return re.sub(r"[^a-z0-9_]", "", s2.lower())


def quote_ident(ident: str) -> sql.Identifier:
    return sql.Identifier("public", ident)


def fetch_public_tables(conn: psycopg.Connection) -> list[str]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema = 'public' AND table_type = 'BASE TABLE'
            ORDER BY table_name
            """
        )
        return [r[0] for r in cur.fetchall()]


def fetch_columns(conn: psycopg.Connection, table: str) -> dict[str, str]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT column_name, data_type
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = %s
            """,
            (table,),
        )
        return {row[0]: row[1] for row in cur.fetchall()}


def choose_link_column(columns: dict[str, str]) -> tuple[str | None, str]:
    if "survey_id" in columns:
        return "survey_id", "survey-linked"
    if "id" in columns and columns["id"].lower() == "uuid":
        return "id", "id-linked"
    return None, "not-survey-linked"


def resolve_table_name(requested: str, existing_tables: set[str]) -> tuple[str | None, str]:
    normalized = normalize_name(requested)

    if requested in existing_tables:
        return requested, "exact"
    if normalized in existing_tables:
        return normalized, "normalized"

    override_key = normalize_name(requested)
    if override_key in TABLE_NAME_OVERRIDES:
        overridden = TABLE_NAME_OVERRIDES[override_key]
        if overridden in existing_tables:
            return overridden, "override"

    close = difflib.get_close_matches(normalized, list(existing_tables), n=1, cutoff=0.90)
    if close:
        return close[0], "fuzzy"

    return None, "not_found"


def expected_records_for(requested: str) -> int:
    normalized = normalize_name(requested)
    for key, value in TABLE_EXPECTED_RECORDS.items():
        if normalize_name(key) == normalized:
            return int(value)
    return 1


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name)
    cleaned = cleaned.strip() or "Sheet"
    return cleaned[:31]


def build_module_resolutions(conn: psycopg.Connection) -> list[TableResolution]:
    existing_tables = set(fetch_public_tables(conn))
    rows: list[TableResolution] = []

    for module, requested_tables in MODULES.items():
        for requested in requested_tables:
            actual, match_type = resolve_table_name(requested, existing_tables)
            if not actual:
                rows.append(
                    TableResolution(
                        module=module,
                        requested_name=requested,
                        actual_table=None,
                        status="NOT_IMPLEMENTED",
                        link_column=None,
                        expected_per_survey=expected_records_for(requested),
                        notes="Table not found in schema",
                    )
                )
                continue

            columns = fetch_columns(conn, actual)
            link_column, link_type = choose_link_column(columns)
            if link_column:
                rows.append(
                    TableResolution(
                        module=module,
                        requested_name=requested,
                        actual_table=actual,
                        status="ACTIVE",
                        link_column=link_column,
                        expected_per_survey=expected_records_for(requested),
                        notes=f"{match_type}; {link_type}",
                    )
                )
            else:
                rows.append(
                    TableResolution(
                        module=module,
                        requested_name=requested,
                        actual_table=actual,
                        status="NOT_IMPLEMENTED",
                        link_column=None,
                        expected_per_survey=expected_records_for(requested),
                        notes=f"{match_type}; no survey_id or uuid id link",
                    )
                )

    return rows


def fetch_district_options(conn: psycopg.Connection) -> list[tuple[str, str]]:
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, district FROM public.districts ORDER BY district"
        )
        return [(r[0], r[1]) for r in cur.fetchall()]


def fetch_year_options(conn: psycopg.Connection) -> list[str]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT DISTINCT survey_year
            FROM public.qr_codes
            WHERE survey_year IS NOT NULL AND survey_year <> ''
            ORDER BY survey_year
            """
        )
        return [r[0] for r in cur.fetchall()]


def fetch_qr_code_type_options(conn: psycopg.Connection) -> list[tuple[str, str]]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id::text, COALESCE(name, '(unnamed)')
            FROM public.qr_code_types
            ORDER BY id
            """
        )
        return [(r[0], r[1]) for r in cur.fetchall()]


def fetch_project_options(conn: psycopg.Connection) -> list[tuple[str, str]]:
    """Fetch active projects where status_id = 1."""
    with conn.cursor() as cur:
        cur.execute(
            """SELECT id, name FROM public.projects WHERE status_id = 1 ORDER BY name"""
        )
        return [(r[0], r[1] or "(unnamed)") for r in cur.fetchall()]


def fetch_health_center_options(
    conn: psycopg.Connection,
    district_ids: list[str] | None = None,
) -> list[tuple[str, str]]:
    """Fetch health centers, optionally filtered to selected district IDs."""
    with conn.cursor() as cur:
        if district_ids:
            cur.execute(
                """
                SELECT hc.id, COALESCE(hc.name, '(unnamed)')
                FROM public.health_centers hc
                WHERE hc.district_id::text = ANY(%s)
                ORDER BY hc.name
                """,
                (district_ids,),
            )
        else:
            cur.execute(
                """
                SELECT hc.id, COALESCE(hc.name, '(unnamed)')
                FROM public.health_centers hc
                ORDER BY hc.name
                """
            )
        return [(r[0], r[1]) for r in cur.fetchall()]


def present_numbered_options(
    items: list[tuple[str, str]], prompt_text: str
) -> list[str]:
    """Present numbered options and return selected IDs based on user input.
    
    Args:
        items: List of (id, display_name) tuples
        prompt_text: Text to display before the list
    
    Returns:
        List of selected IDs from items, or empty list if none selected
    """
    if not items:
        return []
    
    print(f"\n{prompt_text}")
    for i, (item_id, item_name) in enumerate(items, 1):
        print(f"{i}. {item_name}")
    
    selection = input("Enter number(s) separated by commas (Enter to skip): ").strip()
    if not selection:
        return []
    
    selected_ids = []
    try:
        indices = [int(x.strip()) for x in selection.split(",")]
        for idx in indices:
            if 1 <= idx <= len(items):
                selected_ids.append(items[idx - 1][0])
    except ValueError:
        print("Invalid input. Skipping.")
    
    return selected_ids


def maybe_prompt(value: str | None, prompt_text: str, interactive: bool) -> str | None:
    if value is not None or not interactive:
        return value
    entered = input(prompt_text).strip()
    return entered or None




def build_scope_query(
    district_id: str | None,
    district_name: str | None,
    year_filter: str | None,
    type_ids: list[str] | None,
    health_center_ids: list[str] | None,
) -> tuple[str, list[Any]]:
    where: list[str] = []
    params: list[Any] = []

    if district_id:
        where.append("q.district_id = %s")
        params.append(district_id)
    elif district_name:
        where.append("COALESCE(d.district, q.district) = %s")
        params.append(district_name)

    if year_filter:
        if "%" in year_filter:
            where.append("q.survey_year LIKE %s")
            params.append(year_filter)
        elif re.fullmatch(r"\d{4}", year_filter):
            where.append("q.survey_year LIKE %s")
            params.append(f"{year_filter}%")
        else:
            where.append("q.survey_year = %s")
            params.append(year_filter)

    if type_ids:
        where.append("q.type_id::text = ANY(%s)")
        params.append(type_ids)

    if health_center_ids:
        where.append("COALESCE(hc.id, q.health_center_id)::text = ANY(%s)")
        params.append(health_center_ids)

    sql_text = """
        SELECT
            q.id AS survey_id,
            q.survey_year,
            q.type_id::text AS qr_code_type_id,
            qct.name AS qr_code_type_name,
            COALESCE(d.id, q.district_id) AS district_id,
            COALESCE(d.district, q.district) AS district_name,
            COALESCE(hc.id, q.health_center_id) AS health_center_id,
            COALESCE(hc.name, q.health_center) AS health_center_name
        FROM public.qr_codes q
        LEFT JOIN public.qr_code_types qct ON qct.id = q.type_id
        LEFT JOIN public.districts d ON d.id = q.district_id
        LEFT JOIN public.health_centers hc ON hc.id = q.health_center_id
    """

    if where:
        sql_text += " WHERE " + " AND ".join(where)

    sql_text += " ORDER BY district_name, health_center_name, q.survey_year, q.id"
    return sql_text, params


def fetch_scope_rows(
    conn: psycopg.Connection,
    district_id: str | None,
    district_name: str | None,
    year_filter: str | None,
    type_ids: list[str] | None,
    health_center_ids: list[str] | None,
) -> list[dict[str, Any]]:
    query, params = build_scope_query(
        district_id,
        district_name,
        year_filter,
        type_ids,
        health_center_ids,
    )
    with conn.cursor() as cur:
        cur.execute(cast(Any, query), params)
        description = cur.description
        if description is None:
            return []
        cols = [d.name for d in description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


def fetch_counts_for_table(
    conn: psycopg.Connection,
    scope_rows: list[dict[str, Any]],
    table_name: str,
    link_column: str,
) -> dict[str, int]:
    if not scope_rows:
        return {}

    survey_ids = [r["survey_id"] for r in scope_rows]

    stmt = sql.SQL(
        """
        SELECT {link_col}::text AS survey_id, COUNT(*)::int AS cnt
        FROM {tbl}
        WHERE {link_col} = ANY(%s)
        GROUP BY {link_col}
        """
    ).format(
        link_col=sql.Identifier(link_column),
        tbl=quote_ident(table_name),
    )

    with conn.cursor() as cur:
        cur.execute(stmt, (survey_ids,))
        return {row[0]: row[1] for row in cur.fetchall()}


def summarize(
    scope_rows: list[dict[str, Any]],
    resolutions: list[TableResolution],
    conn: psycopg.Connection,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    total_surveys = len(scope_rows)

    table_rows: list[dict[str, Any]] = []
    health_center_rows: list[dict[str, Any]] = []
    module_rollup: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "module": "",
            "tables_total": 0,
            "tables_active": 0,
            "tables_not_implemented": 0,
            "expected_records_total": 0,
            "total_records": 0,
            "missing_records": 0,
            "surveys_expected_total": 0,
            "surveys_with_data_total": 0,
            "surveys_missing_total": 0,
        }
    )

    by_hc: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in scope_rows:
        by_hc[row["health_center_name"] or "Unknown health center"].append(row)

    for res in resolutions:
        counts = {}
        if res.status == "ACTIVE" and res.actual_table and res.link_column:
            counts = fetch_counts_for_table(conn, scope_rows, res.actual_table, res.link_column)

        expected_total = res.expected_per_survey * total_surveys
        total_records = sum(counts.values()) if counts else 0
        missing_records = max(expected_total - total_records, 0)
        surveys_with_data = sum(1 for s in scope_rows if counts.get(str(s["survey_id"]), 0) > 0)
        surveys_missing = total_surveys - surveys_with_data

        table_rows.append(
            {
                "module": res.module,
                "requested_table": res.requested_name,
                "actual_table": res.actual_table or "",
                "status": res.status,
                "link_column": res.link_column or "",
                "expected_per_survey": res.expected_per_survey,
                "surveys_expected": total_surveys,
                "surveys_with_data": surveys_with_data,
                "surveys_missing": surveys_missing,
                "expected_records_total": expected_total,
                "total_records": total_records,
                "missing_records": missing_records,
                "notes": res.notes,
            }
        )

        mod = module_rollup[res.module]
        mod["module"] = res.module
        mod["tables_total"] += 1
        mod["expected_records_total"] += expected_total
        mod["total_records"] += total_records
        mod["missing_records"] += missing_records
        mod["surveys_expected_total"] += total_surveys
        mod["surveys_with_data_total"] += surveys_with_data
        mod["surveys_missing_total"] += surveys_missing
        if res.status == "ACTIVE" and res.actual_table and res.link_column:
            mod["tables_active"] += 1
        else:
            mod["tables_not_implemented"] += 1

        for health_center_name, hc_scope_rows in by_hc.items():
            hc_expected_total = res.expected_per_survey * len(hc_scope_rows)
            hc_total_records = sum(
                counts.get(str(row["survey_id"]), 0) for row in hc_scope_rows
            ) if counts else 0
            hc_missing_records = max(hc_expected_total - hc_total_records, 0)
            hc_surveys_with_data = sum(
                1 for row in hc_scope_rows if counts.get(str(row["survey_id"]), 0) > 0
            ) if counts else 0
            hc_surveys_missing = len(hc_scope_rows) - hc_surveys_with_data

            health_center_rows.append(
                {
                    "module": res.module,
                    "requested_table": res.requested_name,
                    "actual_table": res.actual_table or "",
                    "health_center": health_center_name,
                    "survey_count": len(hc_scope_rows),
                    "expected_per_survey": res.expected_per_survey,
                    "surveys_with_data": hc_surveys_with_data,
                    "surveys_missing": hc_surveys_missing,
                    "expected_records_total": hc_expected_total,
                    "total_records": hc_total_records,
                    "missing_records": hc_missing_records,
                    "status": res.status,
                    "notes": res.notes,
                }
            )

    module_rows = sorted(module_rollup.values(), key=lambda row: row["module"])
    table_rows.sort(key=lambda row: (row["module"], row["requested_table"]))
    health_center_rows.sort(
        key=lambda row: (row["module"], row["requested_table"], row["health_center"])
    )

    return module_rows, table_rows, health_center_rows


def summarize_project_data(
    conn: psycopg.Connection,
    project_ids: list[str],
) -> list[dict[str, Any]]:
    if not project_ids:
        return []

    rows: list[dict[str, Any]] = []

    with conn.cursor() as cur:
        for project_id in project_ids:
            cur.execute(
                "SELECT id, name FROM public.projects WHERE id = %s",
                (project_id,),
            )
            project_row = cur.fetchone()
            exists = project_row is not None
            name = project_row[1] if project_row else ""

            cur.execute(
                """
                SELECT COUNT(*)::int
                FROM public.project_questions_maps
                WHERE project_id = %s
                """,
                (project_id,),
            )
            question_mappings_row = cur.fetchone()
            question_mappings = int(question_mappings_row[0]) if question_mappings_row else 0

            cur.execute(
                """
                SELECT COUNT(DISTINCT qcm.category_id)::int
                FROM public.project_questions_maps pqm
                LEFT JOIN public.question_categories_maps qcm
                  ON qcm.question_id = pqm.question_id
                WHERE pqm.project_id = %s
                """,
                (project_id,),
            )
            categories_count_row = cur.fetchone()
            categories_count = int(categories_count_row[0]) if categories_count_row else 0

            cur.execute(
                """
                SELECT COUNT(*)::int
                FROM public.project_questions_maps pqm
                LEFT JOIN public.project_questions pq
                  ON pq.id = pqm.question_id
                WHERE pqm.project_id = %s AND pq.id IS NULL
                """,
                (project_id,),
            )
            orphan_question_maps_row = cur.fetchone()
            orphan_question_maps = (
                int(orphan_question_maps_row[0]) if orphan_question_maps_row else 0
            )

            rows.append(
                {
                    "project_id": project_id,
                    "project_exists": "yes" if exists else "no",
                    "project_name": name,
                    "question_mappings": question_mappings,
                    "distinct_categories": categories_count,
                    "orphan_question_maps": orphan_question_maps,
                }
            )

    return rows


def get_project_facility_data(
    conn: psycopg.Connection,
    project_ids: list[str],
) -> list[dict[str, Any]]:
    """Get project data by facility: project_id, project_name, health_center, item_count."""
    if not project_ids:
        return []

    rows: list[dict[str, Any]] = []

    with conn.cursor() as cur:
        # Build IN clause for multiple project IDs
        ids_placeholder = ",".join(["%s"] * len(project_ids))
        query = f"""
            SELECT
                pd.project_id,
                p.name as project_name,
                hc.name as health_center,
                COUNT(*)::int as item_count
            FROM public.project_data pd
            LEFT JOIN public.projects p ON p.id = pd.project_id
            LEFT JOIN public.health_centers hc ON hc.id = pd.health_center_id
            WHERE pd.project_id IN ({ids_placeholder})
            GROUP BY pd.project_id, p.name, hc.name
            ORDER BY p.name, hc.name
        """
        cur.execute(cast(Any, query), project_ids)
        result = cur.fetchall()

        if result and result[0] and len(cast(Any, result[0])) >= 4:
            for row in result:
                rows.append(
                    {
                        "Project ID": row[0],
                        "Project Name": row[1] or "(unnamed)",
                        "Health Center": row[2] or "(unknown)",
                        "Items Recorded": row[3],
                    }
                )

    return rows


def write_workbook(
    workbook_path: Path,
    module_rows: list[dict[str, Any]],
    health_center_rows: list[dict[str, Any]],
    project_facility_rows: list[dict[str, Any]] | None = None,
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    header_font = Font(bold=True)

    def write_sheet(title: str, rows: list[dict[str, Any]], is_overview: bool = False) -> tuple[str, bool]:
        sheet_name = safe_sheet_name(title)
        sheet = workbook.create_sheet(title=sheet_name)
        has_data = len(rows) > 0

        if not rows:
            sheet.append(["No data"])
            return sheet_name, False

        headers = list(rows[0].keys())
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            sheet.append([row.get(header, "") for header in headers])

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_width = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_width:
                    max_width = len(value)
            sheet.column_dimensions[column_letter].width = min(max_width + 2, 48)

        return sheet_name, has_data

    overview_rows = [
        {
            "Module": row["module"],
            "Tables Total": row["tables_total"],
            "Tables Active": row["tables_active"],
            "Expected Records": row["expected_records_total"],
            "Total Records": row["total_records"],
            "Missing Records": row["missing_records"],
        }
        for row in module_rows
        if row["missing_records"] > 0
    ]
    overview_name, _ = write_sheet("Overview", overview_rows, is_overview=True)

    modules = sorted({row["module"] for row in health_center_rows})
    sheets_to_hide = []
    for module_name in modules:
        rows = [
            {
                "Health Center": row["health_center"],
                "Table": row["requested_table"],
                "Surveys": row["survey_count"],
                "Expected/Survey": row["expected_per_survey"],
                "With Data": row["surveys_with_data"],
                "Missing": row["surveys_missing"],
                "Expected Total": row["expected_records_total"],
                "Total Records": row["total_records"],
                "Missing Records": row["missing_records"],
            }
            for row in health_center_rows
            if row["module"] == module_name and row["missing_records"] > 0
        ]
        sheet_name, has_data = write_sheet(module_name, rows)
        if not has_data:
            sheets_to_hide.append(sheet_name)

    for sheet_name in sheets_to_hide:
        if sheet_name in workbook.sheetnames:
            workbook[sheet_name].sheet_state = "hidden"

    if project_facility_rows:
        write_sheet("Projects", project_facility_rows)

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)


def _format_table(rows: list[dict[str, Any]], max_rows: int = 20) -> str:
    if not rows:
        return "(no rows)"

    display_rows = rows[:max_rows]
    cols = list(display_rows[0].keys())
    widths = {c: len(c) for c in cols}

    for row in display_rows:
        for col in cols:
            widths[col] = max(widths[col], len(str(row.get(col, ""))))

    line = " | ".join(col.ljust(widths[col]) for col in cols)
    sep = "-+-".join("-" * widths[col] for col in cols)
    body = [
        " | ".join(str(row.get(col, "")).ljust(widths[col]) for col in cols)
        for row in display_rows
    ]

    footer = ""
    if len(rows) > max_rows:
        footer = f"\n... {len(rows) - max_rows} more rows"

    return "\n".join([line, sep, *body]) + footer


def write_markdown_report(
    report_path: Path,
    context: dict[str, Any],
    health_center_rows: list[dict[str, Any]],
    project_rows: list[dict[str, Any]],
) -> None:
    by_health_center: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in health_center_rows:
        if row["missing_records"] > 0:
            by_health_center[row["health_center"]].append(row)

    lines = [
        "# Missing Data Report",
        "",
        f"Generated at: {context['generated_at']}",
        f"District filter: {context['district_filter']}",
        f"Health centers filter: {context.get('health_centers_filter', 'ALL')}",
        f"Year filter: {context['year_filter']}",
        f"QR code types filter: {context.get('qr_code_types_filter', 'ALL')}",
        f"Expected surveys in scope: {context['scope_count']}",
        f"Health centers with missing data: {len(by_health_center)}",
        "",
    ]

    for health_center_name in sorted(by_health_center):
        rows = by_health_center[health_center_name]
        lines.extend([f"## Health center: {health_center_name}", ""])

        rows_by_module: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            rows_by_module[row["module"]].append(row)

        for module_name in sorted(rows_by_module):
            lines.append(f"### {module_name}")
            for row in sorted(rows_by_module[module_name], key=lambda item: item["requested_table"]):
                lines.append(
                    f"- {row['requested_table']} (added:{row['total_records']}/expected:{row['expected_records_total']}; missing:{row['missing_records']})"
                )
            lines.append("")

    lines.extend([
        "## Project Data",
        "",
    ])

    if project_rows:
        for row in project_rows:
            lines.append(
                f"- {row['project_id']} ({row['project_name'] or 'unnamed'}; mappings:{row['question_mappings']}; categories:{row['distinct_categories']}; orphans:{row['orphan_question_maps']})"
            )
    else:
        lines.append("- No project rows")

    report_path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def print_console(
    module_rows: list[dict[str, Any]],
    table_rows: list[dict[str, Any]],
    health_center_rows: list[dict[str, Any]],
    project_rows: list[dict[str, Any]],
) -> None:
    missing_modules = [row for row in module_rows if row["missing_records"] > 0]
    missing_tables = [row for row in table_rows if row["missing_records"] > 0]
    missing_health_centers = [row for row in health_center_rows if row["missing_records"] > 0]

    print("\n=== MODULE SUMMARY (missing only) ===")
    print(_format_table(missing_modules, max_rows=50))

    print("\n=== TABLE SUMMARY (missing only, first 10 rows) ===")
    print(_format_table(missing_tables, max_rows=10))

    print("\n=== HEALTH CENTER SUMMARY (missing only, first 10 rows) ===")
    print(_format_table(missing_health_centers, max_rows=10))

    print("\n=== PROJECT DATA SUMMARY ===")
    print(_format_table(project_rows, max_rows=10))


def validate_project_tables(conn: psycopg.Connection) -> list[str]:
    existing = set(fetch_public_tables(conn))
    return [t for t in PROJECT_DATA_TABLES if t not in existing]


def main() -> int:
    args = parse_args()

    try:
        env = load_env(args.env_file)
    except Exception as exc:
        print(f"Failed to read env file: {exc}", file=sys.stderr)
        return 2

    dsn = db_dsn(env)

    try:
        conn = psycopg.connect(dsn)
    except Exception as exc:
        print("Database connection failed.", file=sys.stderr)
        print(f"Error: {exc}", file=sys.stderr)
        print(
            "Hint: confirm VPN is connected and DB_* values in .env are correct.",
            file=sys.stderr,
        )
        return 3

    with conn:
        interactive = not args.no_interactive

        district_id = args.district_id
        district_name = args.district
        year_filter = args.year
        type_ids = [value.strip() for value in args.type_ids.split(",")] if args.type_ids else []
        project_ids = args.projects.split(",") if args.projects else []
        health_center_ids: list[str] = []
        selected_district_ids: list[str] = [district_id] if district_id else []

        # Interactive district selection with numbered options
        if interactive and not district_id and not district_name:
            district_options = fetch_district_options(conn)
            if district_options:
                selected_district_ids = present_numbered_options(
                    district_options, "Available districts:"
                )
                if selected_district_ids:
                    if len(selected_district_ids) == 1:
                        district_id = selected_district_ids[0]
                        district_name = next(
                            (name for did, name in district_options if did == district_id),
                            None,
                        )

        # Interactive health center selection with numbered options
        # If district(s) are selected, only facilities from those districts are shown.
        if interactive:
            health_center_options = fetch_health_center_options(conn, selected_district_ids)
            if health_center_options:
                selected_health_centers = present_numbered_options(
                    health_center_options,
                    "Available health centers"
                    + (
                        " (in selected district(s)):"
                        if selected_district_ids
                        else ":"
                    ),
                )
                if selected_health_centers:
                    health_center_ids = selected_health_centers

        # Interactive year selection with numbered options
        if interactive and not year_filter:
            year_options = fetch_year_options(conn)
            if year_options:
                year_tuples = [(y, y) for y in year_options]
                selected_years = present_numbered_options(
                    year_tuples, "Available survey years:"
                )
                if selected_years:
                    if len(selected_years) == 1:
                        year_filter = selected_years[0]

        if interactive and not type_ids:
            type_options = fetch_qr_code_type_options(conn)
            if type_options:
                selected_type_ids = present_numbered_options(
                    type_options,
                    "Available qr code types:",
                )
                if selected_type_ids:
                    type_ids = selected_type_ids

        # Interactive project selection with numbered options
        if interactive and not project_ids:
            project_options = fetch_project_options(conn)
            if project_options:
                selected_projects = present_numbered_options(
                    project_options, "Active projects (status_id=1):"
                )
                if selected_projects:
                    project_ids = selected_projects
        
        if not project_ids:
            project_ids = DEFAULT_PROJECT_IDS[:]

        missing_project_tables = validate_project_tables(conn)
        if missing_project_tables:
            print(
                "Warning: some project tables are missing: "
                + ", ".join(missing_project_tables),
                file=sys.stderr,
            )

        resolutions = build_module_resolutions(conn)
        scope_rows = fetch_scope_rows(
            conn,
            district_id,
            district_name,
            year_filter,
            type_ids,
            health_center_ids,
        )

        module_rows, table_rows, health_center_rows = summarize(
            scope_rows, resolutions, conn
        )

        project_rows = summarize_project_data(conn, project_ids)
        project_facility_rows = get_project_facility_data(conn, project_ids)

        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        district_display = (district_name or "all_districts").replace(" ", "_").replace("/", "_")
        run_dir = Path(args.output_dir) / f"{district_display}_{ts}"
        run_dir.mkdir(parents=True, exist_ok=True)

        workbook_path = run_dir / "health_center_summary.xlsx"
        write_workbook(workbook_path, module_rows, health_center_rows, project_facility_rows)

        write_markdown_report(
            run_dir / "report.md",
            {
                "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
                "district_filter": district_id or district_name or "ALL",
                "year_filter": year_filter or "ALL",
                "qr_code_types_filter": ", ".join(type_ids) if type_ids else "ALL",
                "health_centers_filter": ", ".join(health_center_ids)
                if health_center_ids
                else "ALL",
                "scope_count": len(scope_rows),
            },
            health_center_rows,
            project_rows,
        )

        print_console(module_rows, table_rows, health_center_rows, project_rows)

        print("\nArtifacts:")
        print(f"- {workbook_path}")
        print(f"- {run_dir / 'report.md'}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
